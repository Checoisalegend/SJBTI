from pathlib import Path
import json
import sys

import openpyxl


def compact(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: inspect_workbook.py <workbook.xlsx>")

    workbook_path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    payload = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, 12),
            max_col=min(ws.max_column, 12),
            values_only=True,
        ):
            rows.append([compact(cell) for cell in row])

        payload.append(
            {
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "preview": rows,
            }
        )

    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
